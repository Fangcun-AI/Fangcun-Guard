"""Detection result browsing, detail lookup, and spreadsheet export."""

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import and_, cast, or_
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.orm import Session

from database.connection import get_admin_db
from database.models import DetectionResult
from models.responses import DetectionResultResponse, PaginatedResponse
from services.application_request_context import resolve_tenant_application_context
from utils.logger import setup_logger
from utils.url_signature import generate_signed_media_url

logger = setup_logger()
router = APIRouter(tags=["Results"])


def get_current_user_and_application_from_request(request: Request, db: Session):
    context = resolve_tenant_application_context(request, db)
    return context.tenant, context.application_id


def _risk_filter(column, value: Optional[str]):
    if not value:
        return None
    return column != "no_risk" if value == "any_risk" else column == value


def _filters(user, application_id, **options):
    risk_columns = (
        DetectionResult.security_risk_level,
        DetectionResult.compliance_risk_level,
        DetectionResult.data_risk_level,
    )
    filters = [
        or_(
            DetectionResult.application_id == application_id,
            and_(
                DetectionResult.application_id.is_(None),
                DetectionResult.is_direct_model_access == True,
                DetectionResult.tenant_id == str(user.id),
            ),
        )
    ]
    overall = options.get("risk_level")
    if overall == "no_risk":
        filters.append(and_(*(column == "no_risk" for column in risk_columns)))
    elif overall == "any_risk":
        filters.append(or_(*(column != "no_risk" for column in risk_columns)))
    elif overall:
        filters.append(or_(*(column == overall for column in risk_columns)))
    for name, column in (
        ("security_risk_level", DetectionResult.security_risk_level),
        ("compliance_risk_level", DetectionResult.compliance_risk_level),
        ("data_risk_level", DetectionResult.data_risk_level),
    ):
        clause = _risk_filter(column, options.get(name))
        if clause is not None:
            filters.append(clause)
    if category := options.get("category"):
        filters.append(or_(*(
            cast(column, JSONB).contains([category])
            for column in (
                DetectionResult.security_categories,
                DetectionResult.compliance_categories,
                DetectionResult.data_categories,
            )
        )))
    if entity := options.get("data_entity_type"):
        filters.append(cast(DetectionResult.data_categories, JSONB).contains([entity]))
    if start := options.get("start_date"):
        filters.append(DetectionResult.created_at >= f"{start} 00:00:00")
    if end := options.get("end_date"):
        filters.append(DetectionResult.created_at <= f"{end} 23:59:59")
    if text := options.get("content_search"):
        filters.append(DetectionResult.content.like(f"%{text}%"))
    if request_id := options.get("request_id_search"):
        filters.append(DetectionResult.request_id.like(f"%{request_id}%"))
    return filters


def _signed_urls(result) -> list[str]:
    urls = []
    for image_path in getattr(result, "image_paths", []) or []:
        try:
            tenant_id, filename = Path(image_path).parts[-2:]
            urls.append(generate_signed_media_url(tenant_id=tenant_id, filename=filename, expires_in_seconds=86400))
        except Exception as exc:
            logger.error("Failed to sign media path %s: %s", image_path, exc)
    return urls


def _response(result, *, truncate: bool = False) -> DetectionResultResponse:
    content = result.content or ""
    return DetectionResultResponse(
        id=result.id,
        request_id=result.request_id,
        content=f"{content[:200]}..." if truncate and len(content) > 200 else content,
        suggest_action=result.suggest_action,
        suggest_answer=result.suggest_answer,
        hit_keywords=result.hit_keywords,
        created_at=result.created_at,
        ip_address=result.ip_address,
        security_risk_level=result.security_risk_level,
        security_categories=result.security_categories,
        compliance_risk_level=result.compliance_risk_level,
        compliance_categories=result.compliance_categories,
        data_risk_level=getattr(result, "data_risk_level", "no_risk"),
        data_categories=getattr(result, "data_categories", []),
        agent_safety_risk_level=getattr(result, "agent_safety_risk_level", "no_risk"),
        agent_safety_categories=getattr(result, "agent_safety_categories", []),
        hallucination_risk_level=getattr(result, "hallucination_risk_level", "no_risk"),
        hallucination_categories=getattr(result, "hallucination_categories", []),
        groundedness_score=getattr(result, "groundedness_score", None),
        consistency_score=getattr(result, "consistency_score", None),
        has_image=getattr(result, "has_image", False),
        image_count=getattr(result, "image_count", 0),
        image_paths=getattr(result, "image_paths", []),
        image_urls=_signed_urls(result),
        is_direct_model_access=getattr(result, "is_direct_model_access", False),
    )


def _query(db: Session, request: Request, options: dict):
    user, application_id = get_current_user_and_application_from_request(request, db)
    return db.query(DetectionResult).filter(and_(*_filters(user, application_id, **options)))


@router.get("/results")
async def get_detection_results(
    request: Request,
    db: Session = Depends(get_admin_db),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    risk_level: Optional[str] = None,
    security_risk_level: Optional[str] = None,
    compliance_risk_level: Optional[str] = None,
    data_risk_level: Optional[str] = None,
    category: Optional[str] = None,
    data_entity_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    content_search: Optional[str] = None,
    request_id_search: Optional[str] = None,
):
    try:
        options = locals().copy()
        query = _query(db, request, options)
        total = query.count()
        records = query.order_by(DetectionResult.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
        return PaginatedResponse(
            items=[_response(record, truncate=True) for record in records],
            total=total,
            page=page,
            per_page=per_page,
            pages=(total + per_page - 1) // per_page,
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Get detection results error: %s", exc)
        raise HTTPException(status_code=500, detail="Failed to get detection results") from exc


_EXPORT_HEADERS = [
    "Request ID", "Detection Content", "Prompt Attack Risk", "Prompt Attack Categories",
    "Content Compliance Risk", "Content Compliance Categories", "Data Leak Risk",
    "Data Leak Categories", "Suggested Action", "Suggested Answer", "Hit Keywords",
    "Has Image", "Image Count", "IP Address", "Detection Time",
]
_EXPORT_WIDTHS = [30, 50, 20, 30, 20, 30, 20, 30, 15, 50, 30, 12, 12, 15, 20]


def _excel(records) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Detection Results"
    fill, font = PatternFill(start_color="366092", end_color="366092", fill_type="solid"), Font(color="FFFFFF", bold=True)
    for column, (header, width) in enumerate(zip(_EXPORT_HEADERS, _EXPORT_WIDTHS), 1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.fill, cell.font, cell.alignment = fill, font, Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[cell.column_letter].width = width
    for row, result in enumerate(records, 2):
        values = [
            result.request_id, result.content, result.security_risk_level or "no_risk",
            ", ".join(result.security_categories or []), result.compliance_risk_level or "no_risk",
            ", ".join(result.compliance_categories or []), result.data_risk_level or "no_risk",
            ", ".join(result.data_categories or []), result.suggest_action, result.suggest_answer or "",
            ", ".join(result.hit_keywords or []), "Yes" if getattr(result, "has_image", False) else "No",
            getattr(result, "image_count", 0), result.ip_address or "",
            result.created_at.strftime("%Y-%m-%d %H:%M:%S") if result.created_at else "",
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row=row, column=column, value=value)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@router.get("/results/export")
async def export_detection_results(
    request: Request,
    db: Session = Depends(get_admin_db),
    risk_level: Optional[str] = None,
    security_risk_level: Optional[str] = None,
    compliance_risk_level: Optional[str] = None,
    data_risk_level: Optional[str] = None,
    category: Optional[str] = None,
    data_entity_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    content_search: Optional[str] = None,
    request_id_search: Optional[str] = None,
):
    try:
        records = _query(db, request, locals().copy()).order_by(DetectionResult.created_at.desc()).limit(10000).all()
        filename = f"detection_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            _excel(records),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Export detection results error: %s", exc)
        raise HTTPException(status_code=500, detail="Failed to export detection results") from exc


@router.get("/results/{result_id}", response_model=DetectionResultResponse)
async def get_detection_result(result_id: int, request: Request, db: Session = Depends(get_admin_db)):
    user, application_id = get_current_user_and_application_from_request(request, db)
    result = db.query(DetectionResult).filter_by(id=result_id).first()
    if not result:
        raise HTTPException(status_code=404, detail="Detection result not found")
    owned = result.application_id == application_id
    owned_dma = result.application_id is None and result.is_direct_model_access and str(result.tenant_id) == str(user.id)
    if not owned and not owned_dma:
        raise HTTPException(status_code=403, detail="Forbidden")
    return _response(result)
