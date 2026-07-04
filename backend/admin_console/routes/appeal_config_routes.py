"""Application appeal configuration, manual review, and export routes."""

from datetime import datetime
from io import BytesIO
from typing import List, Optional
import uuid

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field
from sqlalchemy import desc
from sqlalchemy.orm import Session

from database.connection import get_admin_db
from database.models import AppealRecord, Application
from services.appeal_service import appeal_service
from services.application_request_context import resolve_admin_application_context
from utils.i18n_loader import get_translation
from utils.logger import setup_logger

logger = setup_logger()
router = APIRouter(prefix="/api/v1/config", tags=["appeal-config"])


def get_current_user_and_application_from_request(request: Request, db: Session) -> tuple:
    context = resolve_admin_application_context(request, db)
    return context.tenant, str(context.application_id)


def get_default_message_template(language: str = "en") -> str:
    return get_translation(language, "appealPage", "defaultMessageTemplate")


class AppealConfigUpdate(BaseModel):
    enabled: bool = False
    message_template: Optional[str] = None
    appeal_base_url: str = ""
    final_reviewer_email: Optional[str] = None


class AppealConfigResponse(AppealConfigUpdate):
    id: Optional[str] = None
    message_template: str
    created_at: Optional[str] = None
    updated_at: Optional[str] = None


class AppealRecordResponse(BaseModel):
    id: str
    request_id: str
    user_id: Optional[str]
    original_content: str
    original_risk_level: str
    original_categories: List[str]
    status: str
    ai_approved: Optional[bool]
    ai_review_result: Optional[str]
    processor_type: Optional[str] = None
    processor_id: Optional[str] = None
    processor_reason: Optional[str] = None
    created_at: Optional[str]
    ai_reviewed_at: Optional[str]
    processed_at: Optional[str] = None


class ManualReviewRequest(BaseModel):
    action: str = Field(pattern="^(approve|reject)$")
    reason: Optional[str] = None


class AppealRecordsListResponse(BaseModel):
    items: List[AppealRecordResponse]
    total: int
    page: int
    page_size: int
    pages: int


def _language(request: Request, default: str = "en") -> str:
    return "zh" if "zh" in request.headers.get("accept-language", default).lower() else "en"


def _failure(label: str, exc: Exception):
    logger.error("%s: %s", label, exc)
    raise HTTPException(status_code=500, detail=f"{label}: {exc}") from exc


@router.get("/appeal", response_model=AppealConfigResponse)
async def get_appeal_config(request: Request, db: Session = Depends(get_admin_db)):
    try:
        _, application_id = get_current_user_and_application_from_request(request, db)
        config = await appeal_service.get_config(application_id, db)
        return AppealConfigResponse(**config) if config else AppealConfigResponse(
            enabled=False, message_template=get_default_message_template(_language(request)), appeal_base_url=""
        )
    except HTTPException:
        raise
    except Exception as exc:
        _failure("Failed to get appeal config", exc)


@router.put("/appeal", response_model=AppealConfigResponse)
async def update_appeal_config(
    config_data: AppealConfigUpdate, request: Request, db: Session = Depends(get_admin_db)
):
    try:
        user, application_id = get_current_user_and_application_from_request(request, db)
        config = await appeal_service.update_config(
            application_id=application_id, tenant_id=str(user.id), config_data=config_data.model_dump(), db=db
        )
        return AppealConfigResponse(**config)
    except HTTPException:
        raise
    except Exception as exc:
        _failure("Failed to update appeal config", exc)


@router.get("/appeal/records", response_model=AppealRecordsListResponse)
async def get_appeal_records(
    request: Request,
    status: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_admin_db),
):
    try:
        _, application_id = get_current_user_and_application_from_request(request, db)
        result = await appeal_service.get_appeal_records(
            application_id=application_id, status=status, page=page, page_size=page_size, db=db
        )
        return AppealRecordsListResponse(**result)
    except HTTPException:
        raise
    except Exception as exc:
        _failure("Failed to get appeal records", exc)


@router.post("/appeal/records/{appeal_id}/review")
async def manual_review_appeal(
    appeal_id: str, review_data: ManualReviewRequest, request: Request, db: Session = Depends(get_admin_db)
):
    try:
        user, _ = get_current_user_and_application_from_request(request, db)
        result = await appeal_service.manual_review_appeal(
            appeal_id=appeal_id,
            action=review_data.action,
            reviewer_email=user.email or "unknown@unknown.com",
            reason=review_data.reason,
            language=_language(request, "zh"),
            db=db,
        )
        if not result.get("success"):
            raise HTTPException(status_code=400, detail=result.get("message", "Manual review failed"))
        return result
    except HTTPException:
        raise
    except Exception as exc:
        _failure("Failed to process manual review", exc)


_HEADERS = [
    "Request ID", "Application", "Appeal User", "Original Content", "Original Risk Level",
    "Original Categories", "Status", "AI Approved", "AI Review Result", "Processor Type",
    "Processor ID", "Processor Reason", "Appeal Time", "AI Review Time", "Process Time",
]
_WIDTHS = [30, 20, 20, 60, 15, 30, 15, 12, 50, 15, 20, 40, 20, 20, 20]
_STATUS = {"pending": "Pending", "reviewing": "Reviewing", "pending_review": "Pending Review", "approved": "Approved", "rejected": "Rejected"}


def _time(value) -> str:
    return value.strftime("%Y-%m-%d %H:%M:%S") if value else "-"


def _excel(records, app_names) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Appeal Records"
    fill, font = PatternFill(start_color="366092", end_color="366092", fill_type="solid"), Font(color="FFFFFF", bold=True)
    for column, (header, width) in enumerate(zip(_HEADERS, _WIDTHS), 1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.fill, cell.font, cell.alignment = fill, font, Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[cell.column_letter].width = width
    for row, record in enumerate(records, 2):
        values = [
            record.request_id, app_names.get(record.application_id, "-"), record.user_id or "-",
            (record.original_content or "")[:500], record.original_risk_level or "-",
            ", ".join(record.original_categories or []), _STATUS.get(record.status, record.status),
            "Yes" if record.ai_approved else ("No" if record.ai_approved is False else "-"),
            record.ai_review_result or "-", record.processor_type or "-", record.processor_id or "-",
            record.processor_reason or "-", _time(record.created_at), _time(record.ai_reviewed_at), _time(record.processed_at),
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row=row, column=column, value=value)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@router.get("/appeal/records/export")
async def export_appeal_records(
    request: Request, status: Optional[str] = None, db: Session = Depends(get_admin_db)
):
    try:
        _, application_id = get_current_user_and_application_from_request(request, db)
        query = db.query(AppealRecord).filter(AppealRecord.application_id == uuid.UUID(application_id))
        if status:
            query = query.filter(AppealRecord.status == status)
        records = query.order_by(desc(AppealRecord.created_at)).limit(10000).all()
        app_ids = {record.application_id for record in records if record.application_id}
        apps = db.query(Application).filter(Application.id.in_(app_ids)).all() if app_ids else []
        filename = f"appeal_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            _excel(records, {app.id: app.name for app in apps}),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Export appeal records error: %s", exc)
        raise HTTPException(status_code=500, detail="Failed to export appeal records") from exc
